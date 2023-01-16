
# You have to install package
        # openpyxl


import time

import openpyxl

# Opening Excel File from Location
worksheet = openpyxl.load_workbook("D:\Screenshot\SS\excel_operations.xlsx")
worksheet_2 = openpyxl.load_workbook("D:\Screenshot\SS\excel.xlsx")
# NOTE: Make sure your file is closed before performing any operation

worksheet.active # ---------------> It will take by default sheet which you have saved list time.
worksheet_1 = worksheet["Ambition"] # -----> We can handle sheet using sheet name.

# reading values from Excel sheet
# row_1 = worksheet_1.cell(row = 1, column= 2).value
# print(row_1)
# row_5 = worksheet_1.cell(row = 5, column= 1).value
# print(row_5)

# Writing values to Excel sheet
# worksheet_1.cell(row = 15, column= 5).value = "Pooja"
# worksheet_1.cell(row = 10, column= 2).value = "Shantanu"
# worksheet.save(filename="D:\Screenshot\SS\excel.xlsx")


for row in range(0, worksheet_1.max_row):
    print(row)
    for col in worksheet_1.iter_cols(1, worksheet_1.max_column):
        print(col[row].value, end= ' ')

"""===================================================================================="""

worksheet_2.active
worksheet_2["sheet1"]

for row in range(0, worksheet_1.max_row):
    print(row)
    for col in worksheet_1.iter_cols(1, worksheet_1.max_column):
        print(col[row].value, end= ' ')

time.sleep(5)

